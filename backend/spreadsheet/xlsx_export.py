"""Build an .xlsx workbook for a sheet, embedding native charts for sparklines (MED-295).

The frontend XLSX export is value-only (SheetJS) and cannot write native Excel
charts. This backend export writes the same cell values plus a native
``openpyxl`` ``LineChart`` for every ``=SPARKLINE(range)`` cell, so the exported
file opens in Excel with a real, editable chart.
"""
import io
import re
from typing import Any, List, Optional

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.utils import get_column_letter
from rest_framework.exceptions import ValidationError

from .models import Cell, ComputedCellType
from .sparkline import is_sparkline, parse_sparkline

# Excel forbids these characters in a worksheet title, and caps it at 31 chars.
_INVALID_TITLE_CHARS = re.compile(r'[\\/?*\[\]:]')


def _safe_sheet_title(name: Optional[str]) -> str:
    """A worksheet title Excel/openpyxl will accept for any user-entered name."""
    cleaned = _INVALID_TITLE_CHARS.sub(' ', (name or '').strip())
    return cleaned[:31] or 'Sheet1'


def _cell_export_value(cell: Cell) -> Optional[Any]:
    """Effective value to write for a (non-sparkline) cell."""
    if cell.computed_type == ComputedCellType.NUMBER and cell.computed_number is not None:
        return float(cell.computed_number)
    if cell.number_value is not None:
        return float(cell.number_value)
    if cell.computed_type == ComputedCellType.STRING and cell.computed_string is not None:
        return cell.computed_string
    if cell.string_value is not None:
        return cell.string_value
    if cell.boolean_value is not None:
        return cell.boolean_value
    return None


def build_sheet_workbook(sheet) -> bytes:
    """Serialize a sheet to .xlsx bytes: cell values + native charts for sparklines."""
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = _safe_sheet_title(sheet.name)

    cells = (
        Cell.objects.filter(
            sheet=sheet, is_deleted=False,
            row__is_deleted=False, column__is_deleted=False,
        )
        .select_related('row', 'column')
    )

    sparkline_cells: List[Cell] = []
    for cell in cells:
        # openpyxl is 1-based; our positions are 0-based.
        row = cell.row.position + 1
        col = cell.column.position + 1
        if is_sparkline(cell.raw_input):
            sparkline_cells.append(cell)
            continue  # the chart is added below; never write its JSON payload
        value = _cell_export_value(cell)
        if value is not None:
            worksheet.cell(row=row, column=col, value=value)

    for cell in sparkline_cells:
        try:
            spec = parse_sparkline(cell.raw_input)
        except ValidationError:
            continue  # a bad spec exports as an empty cell, not a crash
        (r0, c0), (r1, c1) = spec.start, spec.end
        chart = LineChart()
        chart.legend = None
        data = Reference(
            worksheet,
            min_col=c0 + 1, max_col=c1 + 1,
            min_row=r0 + 1, max_row=r1 + 1,
        )
        # A single-row range lays its series out across a row.
        chart.add_data(data, from_rows=(r0 == r1))
        # Carry the spec color onto the native series line (openpyxl wants RRGGBB).
        if spec.color and chart.series:
            chart.series[0].graphicalProperties.line.solidFill = spec.color.lstrip('#')
        anchor = f'{get_column_letter(cell.column.position + 1)}{cell.row.position + 1}'
        worksheet.add_chart(chart, anchor)

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
