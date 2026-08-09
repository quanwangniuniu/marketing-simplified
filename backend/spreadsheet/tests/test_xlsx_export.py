"""XLSX export embeds a native chart for sparkline cells (MED-295)."""
import io
from decimal import Decimal

from django.test import TestCase
from openpyxl import load_workbook
from openpyxl.chart import LineChart
from rest_framework.test import APIClient

from spreadsheet.models import Cell, CellValueType, ComputedCellType, SheetColumn, SheetRow
from spreadsheet.xlsx_export import build_sheet_workbook
from spreadsheet.tests.test_services import (
    create_test_organization, create_test_project, create_test_sheet,
    create_test_spreadsheet, create_test_user,
)


class TestXlsxExport(TestCase):
    def setUp(self):
        self.user = create_test_user()
        self.org = create_test_organization()
        self.project = create_test_project(self.org, owner=self.user)
        self.spreadsheet = create_test_spreadsheet(self.project)
        self.sheet = create_test_sheet(self.spreadsheet)
        self.rows = {i: SheetRow.objects.create(sheet=self.sheet, position=i) for i in range(5)}
        self.cols = {
            i: SheetColumn.objects.create(sheet=self.sheet, position=i, name=chr(ord('A') + i))
            for i in range(5)
        }
        for r, v in enumerate([10, 20, 15, 30, 25]):
            Cell.objects.create(
                sheet=self.sheet, row=self.rows[r], column=self.cols[0],
                value_type=CellValueType.NUMBER, number_value=Decimal(v),
                computed_type=ComputedCellType.NUMBER, computed_number=Decimal(v),
            )
        # sparkline in E1 charting A1:A5
        Cell.objects.create(
            sheet=self.sheet, row=self.rows[0], column=self.cols[4],
            value_type=CellValueType.FORMULA,
            raw_input='=SPARKLINE(A1:A5)', formula_value='=SPARKLINE(A1:A5)',
        )

    def _load(self):
        return load_workbook(io.BytesIO(build_sheet_workbook(self.sheet)))

    def test_values_are_written(self):
        ws = self._load().active
        assert ws['A1'].value == 10
        assert ws['A5'].value == 25

    def test_sparkline_becomes_a_native_line_chart(self):
        ws = self._load().active
        assert len(ws._charts) == 1
        assert isinstance(ws._charts[0], LineChart)

    def test_sparkline_cell_has_no_json_payload(self):
        ws = self._load().active
        assert ws['E1'].value is None  # the JSON payload is never written as a value

    def test_chart_is_linked_to_the_source_range(self):
        # The native chart must reference the real cells (A1:A5), not embed a snapshot.
        ws = self._load().active
        series = ws._charts[0].series
        assert len(series) == 1
        ref = series[0].val.numRef.f  # e.g. "Sheet1!$A$1:$A$5"
        assert '$A$1:$A$5' in ref

    def test_chart_series_carries_spec_color(self):
        # A colored spec paints the native series line (openpyxl stores RRGGBB, no '#').
        Cell.objects.create(
            sheet=self.sheet, row=self.rows[1], column=self.cols[4],
            value_type=CellValueType.FORMULA,
            raw_input='=SPARKLINE(A1:A5, "line", "#3CCED7")',
            formula_value='=SPARKLINE(A1:A5, "line", "#3CCED7")',
        )
        ws = self._load().active
        fills = [
            c.series[0].graphicalProperties.line.solidFill
            for c in ws._charts
            if c.series and c.series[0].graphicalProperties.line.solidFill is not None
        ]
        assert any('3CCED7' in str(f) for f in fills)

    def test_export_survives_illegal_sheet_title(self):
        # A sheet named with Excel-illegal chars must not crash the export.
        self.sheet.name = 'Q1/Q2:[2026]*?'
        self.sheet.save()
        wb = self._load()  # would raise before the title is sanitized
        title = wb.active.title
        assert not any(ch in title for ch in '\\/?*[]:')
        assert len(title) <= 31

    def test_export_endpoint_sanitizes_content_disposition(self):
        # A quote/newline in the sheet name must not break or inject the header.
        self.sheet.name = 'evil"\r\nname'
        self.sheet.save()
        client = APIClient()
        client.force_authenticate(user=self.user)
        url = (
            f'/api/spreadsheet/spreadsheets/{self.spreadsheet.slug}/'
            f'sheets/{self.sheet.id}/export.xlsx'
        )
        response = client.get(url)
        assert response.status_code == 200
        disposition = response['Content-Disposition']
        assert '"' not in disposition.replace('filename="', '').rstrip('"')
        assert '\r' not in disposition and '\n' not in disposition

    def test_export_endpoint_returns_xlsx_with_chart(self):
        client = APIClient()
        client.force_authenticate(user=self.user)
        url = (
            f'/api/spreadsheet/spreadsheets/{self.spreadsheet.slug}/'
            f'sheets/{self.sheet.id}/export.xlsx'
        )
        response = client.get(url)
        assert response.status_code == 200
        assert 'spreadsheetml.sheet' in response['Content-Type']
        assert 'attachment' in response['Content-Disposition']
        ws = load_workbook(io.BytesIO(response.content)).active
        assert len(ws._charts) == 1
        assert isinstance(ws._charts[0], LineChart)
